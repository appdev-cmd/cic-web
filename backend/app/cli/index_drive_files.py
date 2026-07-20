import io
import os
import re
import asyncio
from sqlalchemy import select
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from openpyxl import load_workbook

from app.core.config import settings
from app.db.session import async_session_factory
from app.models.document_upload import Document, DocumentChunk
from app.services.embedding_service import get_embedding
from app.services.document_processor import process_document_to_chunks


def extract_drive_id(url: str) -> tuple[str, str] | None:
    """Trích xuất file_id từ Google Docs hoặc Google Drive URL."""
    if not url:
        return None

    # Ví dụ: https://docs.google.com/document/d/1VuieB2R-DntfoEX8MfJ2pZjxbWYvAC8k/edit...
    doc_match = re.search(r"/document/d/([a-zA-Z0-9-_]+)", url)
    if doc_match:
        return doc_match.group(1), "google-doc"

    # Ví dụ: https://drive.google.com/file/d/1s6UHYoovdn82BTnCA2_Aha-y-dLM1cOL/view...
    file_match = re.search(r"/file/d/([a-zA-Z0-9-_]+)", url)
    if file_match:
        return file_match.group(1), "drive-file"

    return None


async def download_drive_file(drive_service, file_id: str, file_type: str) -> bytes | None:
    """Tải tệp tin về dưới dạng bytes."""
    try:
        if file_type == "google-doc":
            # Nếu là Google Docs gốc, xuất sang Word (.docx)
            print(f"Exporting Google Doc {file_id} to DOCX...")
            request = drive_service.files().export_media(
                fileId=file_id,
                mimeType="application/vnd.openxmlformats-officedocument.wordprocessingml.document"
            )
        else:
            # Nếu là file upload (.pdf, .docx gốc)
            print(f"Downloading Drive File {file_id}...")
            request = drive_service.files().get_media(fileId=file_id)

        file_stream = io.BytesIO()
        downloader = MediaIoBaseDownload(file_stream, request)
        done = False
        while not done:
            status, done = downloader.next_chunk()
        return file_stream.getvalue()
    except Exception as e:
        print(f"Error downloading file {file_id}: {e}")
        return None


async def index_legal_files(max_files: int = 5) -> None:
    """Đọc tệp Excel, lấy các link J/K, tải về, parse OCR/text và lập chỉ mục vào Supabase."""
    if not settings.GOOGLE_SERVICE_ACCOUNT_FILE or not os.path.exists(settings.GOOGLE_SERVICE_ACCOUNT_FILE):
        print(f"Google Service Account file not found.")
        return

    print("Connecting to Google APIs...")
    creds = service_account.Credentials.from_service_account_file(
        settings.GOOGLE_SERVICE_ACCOUNT_FILE,
        scopes=[
            "https://www.googleapis.com/auth/spreadsheets.readonly",
            "https://www.googleapis.com/auth/drive.readonly"
        ]
    )

    drive_service = build("drive", "v3", credentials=creds)

    # 1. Tải và đọc tệp Excel chỉ mục chính
    print("Downloading Excel index file from Drive...")
    request = drive_service.files().get_media(fileId=settings.GOOGLE_SPREADSHEET_ID)
    file_stream = io.BytesIO()
    downloader = MediaIoBaseDownload(file_stream, request)
    done = False
    while not done:
        status, done = downloader.next_chunk()

    wb = load_workbook(io.BytesIO(file_stream.getvalue()), data_only=False)
    sheet_name = settings.GOOGLE_WORKSHEET_NAME
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
    print(f"Read worksheet: {ws.title}")

    processed_count = 0

    # Duyệt từ hàng thứ 3 (bỏ qua tiêu đề)
    for r_idx in range(3, ws.max_row + 1):
        if processed_count >= max_files:
            print(f"Reached max limit of {max_files} files to index this run.")
            break

        # Số hiệu văn bản (Cột E - Index 5)
        sku_val = ws.cell(row=r_idx, column=5).value
        # Nội dung tóm tắt (Cột F - Index 6)
        name_val = ws.cell(row=r_idx, column=6).value

        if not sku_val or not name_val:
            continue

        sku = str(sku_val).strip()
        filename = f"{sku}.docx"

        # Cột J: Link DOCX (Index 10), Cột K: Link PDF (Index 11)
        cell_docx = ws.cell(row=r_idx, column=10)
        cell_pdf = ws.cell(row=r_idx, column=11)

        link_val = None
        file_type = None

        if cell_docx.hyperlink:
            link_val = cell_docx.hyperlink.target
            file_type = "google-doc"
        elif cell_pdf.hyperlink:
            link_val = cell_pdf.hyperlink.target
            file_type = "drive-file"

        if not link_val:
            continue

        parsed = extract_drive_id(link_val)
        if not parsed:
            continue

        file_id, drive_type = parsed
        if drive_type == "google-doc":
            file_type = "google-doc"

        # Đổi tên đuôi file nếu là PDF
        if file_type == "drive-file" and ("pdf" in link_val.lower() or "file" in link_val.lower()):
            filename = f"{sku}.pdf"

        # 2. Kiểm tra tài liệu đã tồn tại trong DB chưa
        async with async_session_factory() as session:
            stmt = select(Document).where(Document.file_path == file_id)
            res = await session.execute(stmt)
            existing = res.scalar_one_or_none()
            if existing and existing.status == "completed":
                print(f"Document {filename} (ID: {file_id}) already indexed. Skipping...")
                continue

        print(f"\nProcessing target document: {filename} (Link: {link_val})")

        # 3. Tải file từ Google Drive
        file_bytes = await download_drive_file(drive_service, file_id, file_type)
        if not file_bytes:
            print(f"Could not download {filename}. Skipping...")
            continue

        # 4. Ghi nhận tài liệu và cập nhật status "processing"
        async with async_session_factory() as session:
            # Lưu file_path là file_id trên Google Drive để nhận diện duy nhất
            doc = Document(
                filename=filename,
                file_path=file_id,
                status="processing"
            )
            session.add(doc)
            await session.commit()
            await session.refresh(doc)
            doc_id = doc.id

        # 5. Phân mảnh văn bản và sinh vector
        try:
            chunks = process_document_to_chunks(filename, file_bytes)
            if not chunks:
                raise ValueError("Không thể trích xuất văn bản từ tệp tin.")

            print(f"Extracted {len(chunks)} chunks from {filename}. Generating embeddings...")

            async with async_session_factory() as session:
                for idx, c in enumerate(chunks):
                    content = c["content"]
                    page_num = c["page_number"]
                    
                    # Trích xuất vector nhúng
                    embedding = await get_embedding(content)
                    
                    chunk = DocumentChunk(
                        document_id=doc_id,
                        content=str(content),
                        page_number=page_num,
                        chunk_index=int(c["chunk_index"]),
                        chapter=c.get("chapter"),
                        article=c.get("article"),
                        clause=c.get("clause"),
                        content_hash=str(c["content_hash"]),
                        embedding=embedding,
                    )
                    session.add(chunk)
                    if idx % 20 == 0:
                        await session.flush()

                # Cập nhật status sang completed
                db_doc = await session.get(Document, doc_id)
                db_doc.status = "completed"
                await session.commit()
                
            print(f"Successfully indexed document: {filename}")
            processed_count += 1

        except Exception as e:
            print(f"Error indexing {filename}: {e}")
            async with async_session_factory() as session:
                db_doc = await session.get(Document, doc_id)
                if db_doc:
                    db_doc.status = "failed"
                    db_doc.error_message = str(e)
                    await session.commit()


async def main():
    print("--- STARTING DRIVE DOCUMENTS INDEXING WORKFLOW ---")
    # Lập chỉ mục tối đa 5 file luật mẫu trước để tránh quá tải API và kiểm định nhanh
    await index_legal_files(max_files=5)
    print("--- WORKFLOW COMPLETED ---")


if __name__ == "__main__":
    asyncio.run(main())
