import asyncio
import os
import io
import re
from sqlalchemy import select
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from openpyxl import load_workbook

from app.core.config import settings
from app.db.session import async_session_factory
from app.models.product import Product
from app.models.document import LegalDocument
from app.models.document_upload import Document, DocumentChunk
from app.services.document_processor import process_document_to_chunks
from app.services.embedding_service import get_embedding
from app.utils.price import parse_price

try:
    GOOGLE_LIBS_AVAILABLE = True
except ImportError:
    GOOGLE_LIBS_AVAILABLE = False


def extract_drive_id(url: str) -> tuple[str, str] | None:
    """Trích xuất file_id từ Google Docs hoặc Google Drive URL."""
    if not url:
        return None

    doc_match = re.search(r"/document/d/([a-zA-Z0-9-_]+)", url)
    if doc_match:
        return doc_match.group(1), "google-doc"

    file_match = re.search(r"/file/d/([a-zA-Z0-9-_]+)", url)
    if file_match:
        return file_match.group(1), "drive-file"

    return None


async def download_drive_file(drive_service, file_id: str, file_type: str) -> bytes | None:
    """Tải tệp tin về từ Google Drive dưới dạng bytes."""
    try:
        if file_type == "google-doc":
            request = drive_service.files().export_media(
                fileId=file_id,
                mimeType="application/vnd.openxmlformats-officedocument.wordprocessingml.document"
            )
        else:
            request = drive_service.files().get_media(fileId=file_id)

        file_stream = io.BytesIO()
        downloader = MediaIoBaseDownload(file_stream, request)
        done = False
        while not done:
            status, done = downloader.next_chunk()
        return file_stream.getvalue()
    except Exception as e:
        print(f"Error downloading file {file_id}: {e}")
        return None


async def sync_products() -> None:
    """Sync products data from Google Sheets to Supabase with upsert handling."""
    if not settings.GOOGLE_SERVICE_ACCOUNT_FILE or not os.path.exists(settings.GOOGLE_SERVICE_ACCOUNT_FILE):
        return

    creds = service_account.Credentials.from_service_account_file(
        settings.GOOGLE_SERVICE_ACCOUNT_FILE,
        scopes=[
            "https://www.googleapis.com/auth/spreadsheets.readonly",
            "https://www.googleapis.com/auth/drive.readonly"
        ]
    )
    service = build("sheets", "v4", credentials=creds)
    sheet = service.spreadsheets()

    range_name = f"{settings.GOOGLE_PRODUCT_WORKSHEET_NAME}!A:N"
    try:
        result = sheet.values().get(spreadsheetId=settings.GOOGLE_PRODUCT_SPREADSHEET_ID, range=range_name).execute()
        values = result.get("values", [])
    except Exception as e:
        print(f"Could not read product sheets: {e}")
        return

    if not values:
        return

    rows = values[1:]

    async with async_session_factory() as session:
        for r in rows:
            if len(r) < 2 or not r[0]:
                continue
            sku = str(r[0]).strip()
            name = str(r[1]).strip()
            category = str(r[2]).strip() if len(r) > 2 and r[2] is not None else ""
            brand = str(r[3]).strip() if len(r) > 3 and r[3] is not None else ""
            description = str(r[4]).strip() if len(r) > 4 and r[4] is not None else ""
            utility = str(r[5]).strip() if len(r) > 5 and r[5] is not None else ""
            specifications = str(r[6]).strip() if len(r) > 6 and r[6] is not None else ""
            standards = str(r[7]).strip() if len(r) > 7 and r[7] is not None else ""
            
            price = parse_price(r[8]) if len(r) > 8 else None
                
            unit = str(r[9]).strip() if len(r) > 9 and r[9] is not None else ""
            
            try:
                stock = int(r[10]) if len(r) > 10 and r[10] is not None else 0
            except ValueError:
                stock = 0
                
            product_link = str(r[11]).strip() if len(r) > 11 and r[11] is not None else ""
            image_link = str(r[12]).strip() if len(r) > 12 and r[12] is not None else ""
            catalogue_link = str(r[13]).strip() if len(r) > 13 and r[13] is not None else ""

            stmt = select(Product).where(Product.sku == sku)
            db_res = await session.execute(stmt)
            existing_product = db_res.scalar_one_or_none()

            # Giá và tồn kho không nằm trong văn bản embedding. Giữ lại vector
            # hiện có để đồng bộ định kỳ không bị chặn bởi quota dịch vụ AI.
            embedding = existing_product.embedding if existing_product else None
            if embedding is None:
                text_to_embed = f"Tên sản phẩm: {name}. Thương hiệu: {brand}. Nhóm sản phẩm: {category}. Mô tả: {description}. Công dụng: {utility}. Thông số kỹ thuật: {specifications}. Tiêu chuẩn: {standards}."
                try:
                    embedding = await get_embedding(text_to_embed)
                except Exception as exc:
                    print(f"Could not embed product {sku}: {exc}")

            if existing_product:
                existing_product.name = name
                existing_product.category = category
                existing_product.brand = brand
                existing_product.description = description
                existing_product.utility = utility
                existing_product.specifications = specifications
                existing_product.standards = standards
                existing_product.price = price
                existing_product.unit = unit
                existing_product.stock = stock
                existing_product.product_link = product_link
                existing_product.image_link = image_link
                existing_product.catalogue_link = catalogue_link
                if embedding is not None:
                    existing_product.embedding = embedding
            else:
                product = Product(
                    sku=sku,
                    name=name,
                    category=category,
                    brand=brand,
                    description=description,
                    utility=utility,
                    specifications=specifications,
                    standards=standards,
                    price=price,
                    unit=unit,
                    stock=stock,
                    product_link=product_link,
                    image_link=image_link,
                    catalogue_link=catalogue_link,
                    embedding=embedding
                )
                session.add(product)

        await session.commit()
        print("Successfully synced products to Supabase.")


async def sync_documents() -> None:
    """Sync legal documents data from Google Sheets/Drive to Supabase, parsing all metadata columns dynamically."""
    if not settings.GOOGLE_SERVICE_ACCOUNT_FILE or not os.path.exists(settings.GOOGLE_SERVICE_ACCOUNT_FILE):
        return

    creds = service_account.Credentials.from_service_account_file(
        settings.GOOGLE_SERVICE_ACCOUNT_FILE,
        scopes=[
            "https://www.googleapis.com/auth/spreadsheets.readonly",
            "https://www.googleapis.com/auth/drive.readonly"
        ]
    )
    
    values = []
    
    # Tải file binary từ Google Drive để parse cả Hyperlinks
    try:
        drive_service = build("drive", "v3", credentials=creds)
        request = drive_service.files().get_media(fileId=settings.GOOGLE_SPREADSHEET_ID)
        
        file_stream = io.BytesIO()
        downloader = MediaIoBaseDownload(file_stream, request)
        done = False
        while not done:
            status, done = downloader.next_chunk()
        
        file_bytes = file_stream.getvalue()
        wb = load_workbook(io.BytesIO(file_bytes), data_only=False)
        sheet_name = settings.GOOGLE_WORKSHEET_NAME
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        
        # Đọc dữ liệu kèm hyperlink
        for row in ws.iter_rows():
            row_vals = []
            for cell in row:
                val = cell.value
                link = cell.hyperlink.target if cell.hyperlink else None
                row_vals.append((val, link))
            values.append(row_vals)
            
        # Không in tên sheet tiếng việt ra terminal để tránh UnicodeEncodeError trên Windows
        print(f"Read Excel sheet successfully containing {len(values)} rows.")
    except Exception as drive_err:
        print(f"Failed to fetch legal documents via Drive API: {drive_err}")
        return

    if not values or len(values) < 2:
        print("No documents found to process.")
        return

    # Lấy tiêu đề các cột ở hàng 1 (Index 1) để map động
    headers = [str(cell[0]).strip().lower() if cell[0] is not None else "" for cell in values[1]]

    # Khớp động các cột metadata
    title_idx = next((i for i, h in enumerate(headers) if "số hiệu" in h), 4)
    content_idx = next((i for i, h in enumerate(headers) if "nội dung" in h), 5)
    
    publish_idx = next((i for i, h in enumerate(headers) if "ban hành" in h), 6)
    effective_idx = next((i for i, h in enumerate(headers) if "hiệu lực" in h), 7)
    expiration_idx = next((i for i, h in enumerate(headers) if "hết hiệu lực" in h or "kết hạn" in h), 8)
    
    link_docx_idx = next((i for i, h in enumerate(headers) if "docx" in h or "word" in h), 9)
    link_pdf_idx = next((i for i, h in enumerate(headers) if "pdf" in h), 10)
    link_eng_idx = next((i for i, h in enumerate(headers) if "english" in h or "tiếng anh" in h), 11)
    notes_idx = next((i for i, h in enumerate(headers) if "ghi chú" in h), 12)
    
    # 2 Cột mới được thêm theo yêu cầu của user
    signer_idx = next((i for i, h in enumerate(headers) if "người ký" in h or "người kí" in h or "xác nhận" in h), -1)
    posted_idx = next((i for i, h in enumerate(headers) if "ngày đăng" in h or "ngày post" in h), -1)

    # Hàng 0: Tiêu đề danh mục, Hàng 1: Tiêu đề cột, Hàng 2: Dòng dữ liệu đầu tiên
    rows = values[2:]

    legal_docs_to_index = []

    async with async_session_factory() as session:
        for idx, r in enumerate(rows):
            # Lọc hàng trống hoặc không đủ cột
            if len(r) <= max(title_idx, content_idx) or r[title_idx][0] is None or r[content_idx][0] is None:
                continue
                
            title = str(r[title_idx][0]).strip()
            content = str(r[content_idx][0]).strip()
            
            if title == "Số hiệu văn bản" or content == "Nội dung":
                continue

            # Xác định Document Type từ cột A, B, C, D
            doc_type = "VĂN BẢN"
            if len(r) > 0 and r[0][0] is not None:
                doc_type = "LUẬT"
            elif len(r) > 1 and r[1][0] is not None:
                doc_type = "NGHỊ ĐỊNH"
            elif len(r) > 2 and r[2][0] is not None:
                doc_type = "THÔNG TƯ"
            elif len(r) > 3 and r[3][0] is not None:
                doc_type = "VĂN BẢN KHÁC"

            publish_date = str(r[publish_idx][0]).strip() if len(r) > publish_idx and r[publish_idx][0] is not None else None
            effective_date = str(r[effective_idx][0]).strip() if len(r) > effective_idx and r[effective_idx][0] is not None else None
            expiration_date = str(r[expiration_idx][0]).strip() if len(r) > expiration_idx and r[expiration_idx][0] is not None else None
            
            link_docx = r[link_docx_idx][1] if len(r) > link_docx_idx else None
            link_pdf = r[link_pdf_idx][1] if len(r) > link_pdf_idx else None
            link_english = r[link_eng_idx][1] if len(r) > link_eng_idx else None
            notes = str(r[notes_idx][0]).strip() if len(r) > notes_idx and r[notes_idx][0] is not None else None
            
            # Đọc 2 cột metadata người ký và ngày đăng mới thêm
            signer = str(r[signer_idx][0]).strip() if signer_idx != -1 and len(r) > signer_idx and r[signer_idx][0] is not None else None
            date_posted = str(r[posted_idx][0]).strip() if posted_idx != -1 and len(r) > posted_idx and r[posted_idx][0] is not None else None

            # Sinh vector nhúng cho mô tả văn bản
            text_to_embed = f"Số hiệu: {title}. Loại văn bản: {doc_type}. Nội dung: {content}."
            embedding = await get_embedding(text_to_embed)

            # Lấy file ID để tải file đính kèm sau này
            target_link = link_docx or link_pdf
            parsed_id = extract_drive_id(target_link) if target_link else None
            drive_file_id = parsed_id[0] if parsed_id else None

            result = await session.execute(
                select(LegalDocument).where(LegalDocument.document_number == title).limit(1)
            )
            doc = result.scalar_one_or_none()
            if doc is None:
                doc = LegalDocument(title=title, content=content, document_number=title)
                session.add(doc)
            doc.title = title
            doc.content = content
            doc.document_type = doc_type
            doc.publish_date = publish_date
            doc.effective_date = effective_date
            doc.expiration_date = expiration_date
            doc.link_docx = link_docx
            doc.link_pdf = link_pdf
            doc.link_english = link_english
            doc.notes = notes
            doc.signer = signer
            doc.date_posted = date_posted
            doc.drive_file_id = drive_file_id
            doc.sheet_row_index = idx + 3
            doc.embedding = embedding
            
            # Lưu lại danh sách các tệp tin cần tải về lập chỉ mục (OCR)
            if target_link and parsed_id:
                legal_docs_to_index.append({
                    "sku": title,
                    "link": target_link,
                    "file_id": parsed_id[0],
                    "file_type": parsed_id[1]
                })

        await session.commit()
        print("Successfully synced legal_documents metadata columns to Supabase.")

    # 4. Tự động tải xuống các file đính kèm (.docx / .pdf) và phân mảnh lập chỉ mục văn bản chi tiết
    print(f"\n--- STARTING DETAILED PDF/DOCX INDEXING PIPELINE ({len(legal_docs_to_index)} FILES FOUND) ---")
    
    # Chỉ lập chỉ mục tối đa 5 file luật trọng tâm của Xây dựng để tránh tràn bộ nhớ và rate limit API
    max_to_process = 5
    processed = 0
    
    for item in legal_docs_to_index:
        if processed >= max_to_process:
            break
            
        file_id = item["file_id"]
        sku = item["sku"]
        file_type = item["file_type"]
        filename = f"{sku}.docx" if file_type == "google-doc" else f"{sku}.pdf"

        # Kiểm tra file đã lập chỉ mục chưa
        async with async_session_factory() as session:
            stmt = select(Document).where(Document.file_path == file_id)
            res = await session.execute(stmt)
            existing = res.scalar_one_or_none()
            if existing and existing.status == "completed":
                print(f"File ID {file_id} already indexed. Skipping...")
                continue

        # Tải file từ Drive
        file_bytes = await download_drive_file(drive_service, file_id, file_type)
        if not file_bytes:
            continue

        # Lưu Document record
        async with async_session_factory() as session:
            db_doc = Document(filename=filename, file_path=file_id, status="processing")
            session.add(db_doc)
            await session.commit()
            await session.refresh(db_doc)
            doc_id = db_doc.id

        try:
            chunks = process_document_to_chunks(filename, file_bytes)
            if chunks:
                async with async_session_factory() as session:
                    for idx, c in enumerate(chunks):
                        c_text = c["content"]
                        page_num = c["page_number"]
                        c_embed = await get_embedding(c_text)
                        
                        db_chunk = DocumentChunk(
                            document_id=doc_id,
                            content=str(c_text),
                            page_number=page_num,
                            chunk_index=int(c["chunk_index"]),
                            chapter=c.get("chapter"),
                            article=c.get("article"),
                            clause=c.get("clause"),
                            content_hash=str(c["content_hash"]),
                            embedding=c_embed,
                        )
                        session.add(db_chunk)
                        if idx % 10 == 0:
                            await session.flush()
                            
                    # Cập nhật status
                    active_doc = await session.get(Document, doc_id)
                    active_doc.status = "completed"
                    await session.commit()
                print(f"Successfully downloaded and indexed full content of file ID {file_id}.")
                processed += 1
            else:
                raise ValueError("Không có text nào được trích xuất.")
        except Exception as e:
            print(f"Error processing document ID {file_id}: {e}")
            async with async_session_factory() as session:
                active_doc = await session.get(Document, doc_id)
                if active_doc:
                    active_doc.status = "failed"
                    active_doc.error_message = str(e)
                    await session.commit()


async def main():
    print("--- STARTING GOOGLE MASTER DATA SYNC ---")
    await sync_products()
    await sync_documents()
    print("--- SYNC COMPLETED ---")


if __name__ == "__main__":
    asyncio.run(main())
