import uuid
from fastapi import APIRouter, Depends, HTTPException, File, UploadFile
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import load_workbook

from app.api.deps import require_roles
from app.db.session import get_db
from app.models.user import User, UserRole
from app.models.product import Product
from app.repositories.product_repository import ProductRepository
from app.schemas.product import ProductCreate, ProductOut, ProductSearchResponse, ProductUpdate
from app.services.embedding_service import get_embedding
from app.utils.price import parse_price
from app.core.features.dependencies import require_feature

router = APIRouter(prefix="/products", tags=["products"])


def _embedding_text(product: Product) -> str:
    return " ".join(filter(None, (product.name, product.brand, product.category, product.description, product.utility, product.use_case, product.specifications)))


@router.post("/manual", response_model=ProductOut, status_code=201)
async def create_product(payload: ProductCreate, session: AsyncSession = Depends(get_db), _: User = Depends(require_roles(UserRole.ADMIN)), __: User = Depends(require_feature("product_manual_crud"))):
    repo = ProductRepository(session)
    if await repo.get_by_sku(payload.sku):
        raise HTTPException(409, detail="SKU đã tồn tại")
    product = Product(**payload.model_dump())
    try: product.embedding = await get_embedding(_embedding_text(product), max_retries=1)
    except Exception: product.embedding = None
    session.add(product); await session.commit(); await session.refresh(product)
    return product


@router.patch("/manual/{product_id}", response_model=ProductOut)
async def update_product(product_id: uuid.UUID, payload: ProductUpdate, session: AsyncSession = Depends(get_db), _: User = Depends(require_roles(UserRole.ADMIN)), __: User = Depends(require_feature("product_manual_crud"))):
    product = await session.get(Product, product_id)
    if product is None: raise HTTPException(404, detail="Không tìm thấy sản phẩm")
    for key, value in payload.model_dump(exclude_unset=True).items(): setattr(product, key, value)
    try: product.embedding = await get_embedding(_embedding_text(product), max_retries=1)
    except Exception: pass
    await session.commit(); await session.refresh(product); return product


@router.delete("/manual/{product_id}", status_code=204)
async def delete_product(product_id: uuid.UUID, session: AsyncSession = Depends(get_db), _: User = Depends(require_roles(UserRole.ADMIN)), __: User = Depends(require_feature("product_manual_crud"))):
    product = await session.get(Product, product_id)
    if product is None: raise HTTPException(404, detail="Không tìm thấy sản phẩm")
    await session.delete(product); await session.commit()


@router.get("/", response_model=list[ProductOut])
async def list_products(
    skip: int = 0,
    limit: int = 50,
    category: str | None = None,
    session: AsyncSession = Depends(get_db),
    _: User = Depends(require_roles(UserRole.ADMIN, UserRole.MANAGER)),
) -> list[Product]:
    repo = ProductRepository(session)
    return await repo.list_products(skip=skip, limit=limit, category=category)


@router.post("/import")
async def import_products(
    file: UploadFile = File(...),
    session: AsyncSession = Depends(get_db),
    _: User = Depends(require_roles(UserRole.ADMIN)),
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(
            status_code=400, detail="Chỉ chấp nhận tệp Excel (.xlsx, .xls)"
        )

    try:
        wb = load_workbook(file.file, read_only=True)
        sheet = wb.active
        rows = list(sheet.iter_rows(values_only=True))
    except Exception as e:
        raise HTTPException(
            status_code=400, detail=f"Không thể đọc tệp Excel: {str(e)}"
        )

    if not rows:
        raise HTTPException(status_code=400, detail="Tệp Excel trống")

    # Giả định hàng đầu tiên là tiêu đề (header)
    data_rows = rows[1:]

    imported = 0
    updated = 0

    repo = ProductRepository(session)
    for r in data_rows:
        # Bỏ qua hàng trống hoặc thiếu thông tin tối thiểu (sku, name)
        if not r or len(r) < 2 or r[0] is None or r[1] is None:
            continue

        sku = str(r[0]).strip()
        name = str(r[1]).strip()
        category = str(r[2]).strip() if len(r) > 2 and r[2] is not None else None
        brand = str(r[3]).strip() if len(r) > 3 and r[3] is not None else None
        description = str(r[4]).strip() if len(r) > 4 and r[4] is not None else None
        utility = str(r[5]).strip() if len(r) > 5 and r[5] is not None else None
        specifications = str(r[6]).strip() if len(r) > 6 and r[6] is not None else None
        standards = str(r[7]).strip() if len(r) > 7 and r[7] is not None else None

        price = parse_price(r[8]) if len(r) > 8 else None

        unit = str(r[9]).strip() if len(r) > 9 and r[9] is not None else None

        stock = None
        if len(r) > 10 and r[10] is not None:
            try:
                stock = int(r[10])
            except ValueError:
                pass

        product_link = str(r[11]).strip() if len(r) > 11 and r[11] is not None else None
        image_link = str(r[12]).strip() if len(r) > 12 and r[12] is not None else None
        catalogue_link = str(r[13]).strip() if len(r) > 13 and r[13] is not None else None

        # Sinh vector embedding cho sản phẩm (gộp đầy đủ metadata)
        text_to_embed = f"Tên sản phẩm: {name}. Thương hiệu: {brand or ''}. Nhóm sản phẩm: {category or ''}. Mô tả: {description or ''}. Công dụng: {utility or ''}. Thông số kỹ thuật: {specifications or ''}. Tiêu chuẩn: {standards or ''}."
        embedding = await get_embedding(text_to_embed)

        existing = await repo.get_by_sku(sku)
        if existing:
            existing.name = name
            existing.category = category
            existing.brand = brand
            existing.description = description
            existing.utility = utility
            existing.specifications = specifications
            existing.standards = standards
            existing.price = price
            existing.unit = unit
            existing.stock = stock
            existing.product_link = product_link
            existing.image_link = image_link
            existing.catalogue_link = catalogue_link
            if embedding:
                existing.embedding = embedding
            updated += 1
        else:
            p = Product(
                sku=sku,
                name=name,
                category=category,
                brand=brand,
                description=description,
                utility=utility,
                specifications=specifications,
                standards=standards,
                price=price,
                unit=unit,
                stock=stock,
                product_link=product_link,
                image_link=image_link,
                catalogue_link=catalogue_link,
                embedding=embedding,
            )
            await repo.create_product(p)
            imported += 1

    await session.commit()
    return {
        "status": "success",
        "imported": imported,
        "updated": updated,
        "message": f"Nhập thành công: tạo mới {imported}, cập nhật {updated}.",
    }


@router.get("/search", response_model=list[ProductSearchResponse])
async def search_products(
    query: str,
    limit: int = 10,
    session: AsyncSession = Depends(get_db),
    _: User = Depends(require_roles(UserRole.ADMIN, UserRole.MANAGER)),
) -> list[ProductSearchResponse]:
    if not query.strip():
        raise HTTPException(status_code=400, detail="Từ khóa tìm kiếm trống")

    repo = ProductRepository(session)

    # 1. Sinh vector nhúng cho câu hỏi
    try:
        query_embedding = await get_embedding(query, max_retries=1)
    except Exception:
        query_embedding = None

    # 2. Tìm kiếm từ khóa (Keyword Search)
    kw_results = await repo.search_by_keyword(query, limit=limit * 2)

    # 3. Tìm kiếm ngữ nghĩa (Semantic Search)
    vector_results = []
    if query_embedding:
        vector_results = await repo.search_by_semantic(query_embedding, limit=limit * 2)

    # 4. Thuật toán Reciprocal Rank Fusion (RRF) để kết hợp kết quả
    k = 60
    rrf_scores = {}
    products_map = {}

    for rank, p in enumerate(kw_results):
        pid = p.id
        products_map[pid] = p
        rrf_scores[pid] = rrf_scores.get(pid, 0.0) + (1.0 / (k + (rank + 1)))

    for rank, (p, distance) in enumerate(vector_results):
        pid = p.id
        products_map[pid] = p
        # RRF cộng thêm điểm dựa trên thứ hạng tìm kiếm ngữ nghĩa
        rrf_scores[pid] = rrf_scores.get(pid, 0.0) + (1.0 / (k + (rank + 1)))

    # Sắp xếp các kết quả theo điểm RRF giảm dần
    sorted_pids = sorted(rrf_scores.keys(), key=lambda x: rrf_scores[x], reverse=True)

    search_responses = []
    for pid in sorted_pids[:limit]:
        search_responses.append(
            ProductSearchResponse(product=ProductOut.model_validate(products_map[pid]), score=rrf_scores[pid])
        )

    return search_responses
